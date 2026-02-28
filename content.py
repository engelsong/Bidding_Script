from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.page import PageMargins


class Content:
    """
    通过 project 实例创建投标文件目录和报价表
    """

    # 公共样式
    title_font = Font(name='宋体', size=24, bold=True)
    header_font = Font(name='仿宋_GB2312', size=14, bold=True)
    normal_font = Font(name='仿宋_GB2312', size=14)
    sub_font = Font(name='仿宋_GB2312', size=12)
    header_border = Border(bottom=Side(style='medium'))
    normal_border = Border(bottom=Side(style='thin', color='80969696'))
    thin_border = Border(
        left=Side(style='thin', color='80969696'),
        right=Side(style='thin', color='80969696'),
        top=Side(style='thin', color='80969696'),
        bottom=Side(style='thin', color='80969696')
    )
    ctr_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    margin = PageMargins(left=0.75, right=0.75, top=1.0, bottom=1.0, header=0.5, footer=0.5)
    cn_numbers = [
        '一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
        '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九', '二十'
    ]

    def __init__(self, project) -> None:
        self.project = project

    @staticmethod
    def _safe_name(name: str) -> str:
        invalid = '<>:"/\\|?*'
        result = name
        for ch in invalid:
            result = result.replace(ch, '_')
        return result.strip()

    def _section_rows(self):
        rows = []
        rows.append(('section', '投标函部分'))
        rows.extend([
            ('entry', '一', '投标函'),
            ('entry', '二', '授权委托书'),
            ('entry', '三', '开标一览表'),
            ('entry', '四', '采购和实施守法廉洁承诺书'),
            ('entry', '五', '项目经理授权和承诺书'),
            ('entry', '六', '企业内控承诺书'),
            ('entry', '七', '未使用投标咨询/投标代理声明函'),
            ('entry', '八', '发挥党组织、纪检组织作用引领保障做好项目实施工作承诺'),
        ])

        rows.append(('section', '技术标部分'))
        rows.extend([
            ('entry', '一', '合同条款偏离表'),
            ('entry', '二', '采购需求偏离表'),
            ('entry', '三', '物资选型文件'),
        ])

        seq = sorted(self.project.commodities.keys())
        for index in seq:
            item = self.project.commodities[index]
            rows.append(('sub', index, item[0]))

        tech_items = [
            '质量保证声明',
            '包装方案',
            '运输计划',
            '自检验收方案',
            '第三方检验方案',
            '对外实施工作主体责任落实承诺书'
        ]
        if self.project.is_tech:
            tech_items.append('技术服务承诺')
        if self.project.is_qa:
            tech_items.append('售后服务承诺')
        if self.project.is_cc:
            tech_items.append('来华培训和接待承诺')
        tech_items.extend([
            '舆情应对方案',
            '风险防范化解方案',
            '主要标的三体系一览表',
        ])
        for i, item in enumerate(tech_items):
            label = self.cn_numbers[3 + i] if 3 + i < len(self.cn_numbers) else str(4 + i)
            rows.append(('entry', label, item))

        rows.append(('section', '经济标部分'))
        rows.extend([
            ('entry', '一', '投标报价总表'),
            ('entry', '二', '物资对内分项报价表'),
            ('entry', '三', '增值税和消费税退抵税额表'),
            ('subplain', '', '增值税、消费税不退不抵承诺书'),
            ('entry', '四', '技术服务费报价表')
        ])

        rows.append(('section', '商务标部分'))
        rows.extend([
            ('entry', '一', '主要标的的同类物资出口业绩一览表'),
            ('entry', '二', '物资选用节能环保产品'),
            ('subplain', '（一）', '物资节能产品一览表'),
            ('subplain', '（二）', '物资环境标志产品一览表'),
            ('entry', '三', '取得质量管理体系认证等声明'),
        ])
        return rows

    def generate_content(self, filename=None):
        wb = Workbook()
        ws = wb.active
        ws.title = '总'

        ws.column_dimensions['A'].width = 13.125
        ws.column_dimensions['B'].width = 58.125
        ws.column_dimensions['C'].width = 18.25

        ws.merge_cells('A1:C1')
        ws['A1'] = '目  录'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.ctr_alignment
        ws.row_dimensions[1].height = 31.5

        ws['A2'] = '序号'
        ws['B2'] = '内容'
        ws['C2'] = '页码'
        ws.row_dimensions[2].height = 18.75
        for col in ('A', 'B', 'C'):
            cell = ws[f'{col}2']
            cell.font = self.header_font
            cell.alignment = self.ctr_alignment
            cell.border = self.normal_border

        row = 3
        for item in self._section_rows():
            kind = item[0]
            ws.row_dimensions[row].height = 18.75

            if kind == 'section':
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = item[1]
                ws[f'A{row}'].font = self.header_font
                ws[f'A{row}'].alignment = self.ctr_alignment
                ws[f'A{row}'].border = self.normal_border
                ws[f'C{row}'].border = self.normal_border
            elif kind == 'entry':
                ws[f'A{row}'] = item[1]
                ws[f'B{row}'] = item[2]
                ws[f'A{row}'].font = self.normal_font
                ws[f'A{row}'].alignment = self.ctr_alignment
                ws[f'B{row}'].font = self.normal_font
                ws[f'B{row}'].alignment = Alignment(vertical='center', wrap_text=True)
                ws[f'C{row}'].alignment = self.ctr_alignment
                for col in ('A', 'B', 'C'):
                    ws[f'{col}{row}'].border = self.normal_border
            elif kind == 'sub':
                ws[f'A{row}'] = item[1]
                ws[f'B{row}'] = item[2]
                ws[f'A{row}'].font = self.sub_font
                ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                ws[f'B{row}'].font = self.sub_font
                ws[f'B{row}'].alignment = Alignment(vertical='center', wrap_text=True)
                ws[f'C{row}'].alignment = self.ctr_alignment
                ws.row_dimensions[row].height = 20.1
                for col in ('A', 'B', 'C'):
                    ws[f'{col}{row}'].border = self.normal_border
            else:
                ws[f'A{row}'] = item[1]
                ws[f'B{row}'] = item[2]
                ws[f'A{row}'].font = self.normal_font
                ws[f'A{row}'].alignment = self.ctr_alignment
                ws[f'B{row}'].font = self.sub_font if item[1] == '' else self.normal_font
                ws[f'B{row}'].alignment = self.left_alignment
                ws[f'C{row}'].alignment = self.ctr_alignment
                for col in ('A', 'B', 'C'):
                    ws[f'{col}{row}'].border = self.normal_border

            row += 1

        ws.print_area = f'A1:C{row - 1}'
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins = self.margin

        ws2 = wb.create_sheet('资格后审')
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 60
        ws2.column_dimensions['C'].width = 10

        ws2.merge_cells('A1:C1')
        ws2['A1'] = '目  录'
        ws2['A1'].font = self.title_font
        ws2['A1'].alignment = self.ctr_alignment
        ws2.row_dimensions[1].height = 50.1

        ws2['A2'] = '序号'
        ws2['B2'] = '内容'
        ws2['C2'] = '页码'
        for col in ('A', 'B', 'C'):
            cell = ws2[f'{col}2']
            cell.font = self.header_font
            cell.alignment = self.ctr_alignment
            cell.border = self.header_border
        ws2.row_dimensions[2].height = 45

        review_rows = [
            (1, '满足《中华人民共和国政府采购法》第二十二条规定及法律法规的其他规定', 1, 14),
            ('', '1-1 投标人资格声明书', 1, 12),
            ('', '1-2 投标人营业执照', 2, 12),
            (2, '具备援外物资项目实施企业资格', 3, 14),
            (3, '投标保证金银行保函', 5, 14),
        ]

        r = 3
        for no, content, page, font_size in review_rows:
            ws2[f'A{r}'] = no if no != '' else None
            ws2[f'B{r}'] = content
            ws2[f'C{r}'] = page
            ws2[f'A{r}'].font = self.normal_font
            ws2[f'C{r}'].font = self.normal_font
            ws2[f'A{r}'].alignment = self.ctr_alignment
            ws2[f'C{r}'].alignment = self.ctr_alignment
            ws2[f'B{r}'].font = Font(name='仿宋_GB2312', size=font_size)
            ws2[f'B{r}'].alignment = self.left_alignment if font_size == 12 else Alignment(vertical='center', wrap_text=True)
            ws2.row_dimensions[r].height = 45
            for col in ('A', 'B', 'C'):
                ws2[f'{col}{r}'].border = self.normal_border
            r += 1

        ws2.print_area = f'A1:C{r - 1}'
        ws2.page_setup.orientation = 'portrait'
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
        ws2.page_margins = PageMargins(left=0.75, right=0.75, top=0.5, bottom=0.5, header=0.1, footer=0.1)

        if not filename:
            filename = f'content-{self._safe_name(self.project.name)}.xlsx'
        wb.save(filename)
        return filename



        if not filename:
            filename = f'quotation-{self._safe_name(self.project.name)}.xlsx'
        wb.save(filename)
        return filename
