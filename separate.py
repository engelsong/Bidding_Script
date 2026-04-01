import re
import subprocess
from os import listdir
from os.path import abspath
from typing import List, Optional

from openpyxl import load_workbook


class Separate:
    """将报价表按工作表拆分为单独文件，并固化为缓存数值。"""

    workbook_pattern = re.compile(r"^投标报价表-?[\w\S]*\.xlsx$")
    sheet_pattern = re.compile(r"^[0-9]{1,2}\.\w*")

    @staticmethod
    def _safe_name(name: str) -> str:
        invalid = '<>:"/\\|?*'
        result = name or "sheet"
        for ch in invalid:
            result = result.replace(ch, "_")
        return result.strip()

    def _find_workbook_name(self) -> str:
        for doc in listdir():
            if re.match(self.workbook_pattern, doc):
                return doc
        raise FileNotFoundError("No 投标报价表-*.xlsx file found in current directory.")

    @staticmethod
    def _refresh_formula_cache(filename: str) -> None:
        workbook_path = abspath(filename).replace("'", "''")
        command = (
            "$excel = $null; "
            "$workbook = $null; "
            "try { "
            "$excel = New-Object -ComObject Excel.Application; "
            "$excel.Visible = $false; "
            "$excel.DisplayAlerts = $false; "
            f"$workbook = $excel.Workbooks.Open('{workbook_path}'); "
            "$excel.CalculateFullRebuild(); "
            "$workbook.Save(); "
            "} finally { "
            "if ($workbook -ne $null) { $workbook.Close($true) }; "
            "if ($excel -ne $null) { $excel.Quit() }"
            " }"
        )
        try:
            subprocess.run(
                ["powershell", "-NoProfile", "-Command", command],
                check=True,
                capture_output=True,
                text=True,
            )
        except (FileNotFoundError, subprocess.CalledProcessError):
            # 未安装 Excel 或当前环境不支持 COM 时，退回到已有缓存值。
            pass

    def generate(self, workbook_name: Optional[str] = None) -> List[str]:
        filename = workbook_name or self._find_workbook_name()
        self._refresh_formula_cache(filename)

        workbook = load_workbook(filename, data_only=True)
        sheet_names = [
            sheet.title
            for sheet in workbook.worksheets
            if re.match(self.sheet_pattern, sheet.title)
        ]
        workbook.close()

        output_files = []
        for name in sheet_names:
            wb_now = load_workbook(filename, data_only=True)
            for sheet in list(wb_now.worksheets):
                if sheet.title != name:
                    wb_now.remove(sheet)

            output_name = f"{self._safe_name(name)}.xlsx"
            wb_now.save(output_name)
            wb_now.close()
            output_files.append(output_name)

        return output_files
