from __future__ import annotations

import re
from datetime import date, datetime
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side



class Quotation:
    """通过 project 实例创建报价表（不依赖模板）。"""

    def __init__(self, project) -> None:
        self.project = project
        self._all_suppliers_last_row = 1
        self._total_sheet_total_row = 9

        self.title_font = Font(name="宋体", size=16, bold=True)
        self.header_font = Font(name="宋体", size=12, bold=True)
        self.normal_font = Font(name="宋体", size=12)
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="000000")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.diag_border = Border(left=thin, right=thin, top=thin, bottom=thin, diagonal=thin, diagonalUp=True)
        self.import_fill = PatternFill("solid", fgColor="FFFF00")
        self.header_fill = PatternFill("solid", fgColor="808080")

    @staticmethod
    def _safe_name(name: str) -> str:  # 用于生成文件名，过滤掉无法使用的字符
        invalid = '<>:"/\\|?*'
        result = name or "project"
        for ch in invalid:
            result = result.replace(ch, "_")
        return result.strip()

    @staticmethod
    def _parse_date(raw) -> date:  # 生成日期
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        if isinstance(raw, str):
            m = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", raw)
            if m:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return date.today()

    @staticmethod
    def _parse_quantity(raw) -> Optional[int]:  # 用于统计物资数量，将字符串转化为int
        if raw is None:
            return 1
        if isinstance(raw, int):
            return raw
        if isinstance(raw, float):
            return int(raw)
        text = str(raw).strip().replace(",", "")
        m = re.search(r"-?\d+", text)
        if not m:
            return None
        return int(m.group(0))

    def _set_columns(self, ws, widths: Dict[str, float]) -> None:  #设置当前worksheet的列宽
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def _style_row(self, ws, row: int, col_start: int, col_end: int, header: bool = False) -> None:
        #  用于修改指定worksheet行号的从几列到几列的样式
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            cell.font = self.header_font if header else self.normal_font
            cell.alignment = self.center
            cell.border = self.border
            # if header:
            #     cell.fill = self.header_fill

    def _build_all_suppliers(self, ws, items: Dict) -> None:
        ws.title = "全部厂家备用"
        self._set_columns(
            ws,
            {
                "A": 7.6,
                "B": 13.1,
                "C": 12.2,
                "D": 10.6,
                "E": 9.8,
                "F": 42.0,
                "G": 12.2,
                "H": 10.0,
                "I": 14.1,
                "J": 13.1,
                "K": 16.2,
                "L": 8.2,
                "M": 7.7,
                "N": 9.6,
                "O": 12.8,
                "P": 12.1,
                "Q": 8.1,
                "R": 11.2,
                "S": 10.0,
                "T": 13.0,
                "U": 13.0,
                "V": 13.0,
                "W": 13.0,
                "X": 13.0,
                "Y": 13.0,
                "Z": 13.0,
                "AA": 13.0,
                "AB": 6.0,
                "AC": 13.0,
                "AD": 15.3,
                "AE": 13.0,
                "AF": 15.6,
                "AG": 14.2,
                "AH": 13.6,
                "AI": 16.7,
                "AJ": 13.5,
            },
        )
        headers = [
            "序号",
            "品名",
            "HS编码",
            "单位",
            "数量",
            "规格",
            "检验标准",
            "品牌",
            "型号",
            "单价",
            "总价",
            "生产厂商",
            "供货商",
            "偏离情况",
            "强制性认证取证情况",
            "生产供货地",
            "交货期",
            "质量体系",
            "有效期",
            "环境体系",
            "有效期",
            "职业健康",
            "有效期",
            "节能",
            "有效期",
            "环境标志",
            "有效期",
            "备注",
            "",
            "排名",
            "性价比",
            "得分",
            "性能",
            "质量保证",
            "三体系",
            "售后",
        ]
        ws.append(headers)
        self._style_row(ws, 1, 1, len(headers), header=True)

        keys = sorted(items.keys())
        for key in keys:
            idx = key + 1
            ws[f"A{idx}"] = items[key][-1]
            ws[f"B{idx}"] = items[key][0]
            ws[f"C{idx}"] = items[key][1]
            ws[f"D{idx}"] = items[key][2]
            ws[f"E{idx}"] = self._parse_quantity(items[key][3])
            ws[f"F{idx}"] = items[key][4]
            ws[f"G{idx}"] = items[key][5]
            ws[f"H{idx}"] = ""
            ws[f"I{idx}"] = ""
            ws[f"J{idx}"] = 1
            ws[f"J{idx}"].number_format = '#,##0.00'
            ws[f"K{idx}"] = f"=J{idx}*E{idx}"
            ws[f"K{idx}"].number_format = '#,##0.00'
            ws[f"L{idx}"] = ""
            ws[f"M{idx}"] = ""
            ws[f"N{idx}"] = "无"
            ws[f"O{idx}"] = "无"
            ws[f"P{idx}"] = ""
            ws[f"Q{idx}"] = ""
            ws[f"R{idx}"] = ""
            ws[f"S{idx}"] = ""
            ws[f"T{idx}"] = ""
            ws[f"U{idx}"] = ""
            ws[f"V{idx}"] = ""
            ws[f"W{idx}"] = ""
            ws[f"X{idx}"] = ""
            ws[f"Y{idx}"] = ""
            ws[f"Z{idx}"] = ""
            ws[f"AA{idx}"] = ""
            ws[f"AB{idx}"] = "-"
            ws[f"AC{idx}"] = ""
            ws[f"AD{idx}"] = 1
            ws[f"AE{idx}"] = ""
            ws[f"AF{idx}"] = ""
            ws[f"AG{idx}"] = ""
            ws[f"AH{idx}"] = ""
            ws[f"AI{idx}"] = ""
            ws[f"AJ{idx}"] = ""
            ws[f"AL{idx}"] = f"=A{idx}&B{idx}&AD{idx}"
            self._style_row(ws, idx, 1, len(headers), header=False)
            ws[f"F{idx}"].alignment = self.left
            ws.row_dimensions[idx].height = max(24, min(120, (str(items[key][4]).count("\n") + 1) * 16))
        self._all_suppliers_last_row = (max(keys) + 1) if keys else 1


    def _build_selector(self, ws, items: Dict) -> None:
        ws.title = "物资选择"
        self._set_columns(ws, {"A": 8, "B": 28, "C": 10})
        keys = sorted(items.keys())
        row_num = len(keys) + 5
        for key in keys:
            ws[f"A{key}"] = items[key][-1]
            ws[f"B{key}"] = items[key][0]
            ws[f"C{key}"] = f'=MATCH(A{key}&B{key}&1,全部厂家备用!AL$1:AL${row_num},0)'
            self._style_row(ws, key, 1, 3, header=False)


    def _build_fee_input(self, ws, items: Dict) -> None:
        ws.title = "运输费用"
        row_num = len(items) + 3

        self._set_columns(
            ws,
            {
                "A": 12,
                "B": 12,
                "C": 14,
                "D": 10,
                "E": 18,
                "F": 6,
                "G": 12,
                "H": 12,
                "I": 14,
                "J": 10,
                "K": 18,
                "L": 6,
                "M": 6,
                "N": 20,
                "O": 10,
                "P": 10,
                "T": 14,
                "U": 16,
                "V": 14,
                "W": 10,
            },
        )

        ws.row_dimensions[1].height = 30
        for r in range(2, row_num + 1):
            ws.row_dimensions[r].height = 26


        merged_ranges = [
            "A1:E1",
            "G1:K1",
            "M1:W1",
            "B2:C2",
            "H2:I2",
            "A3:A5",
            "A6:A8",
            "A9:A11",
            "A16:A17",
            "A18:D18",
            "G3:G5",
            "G6:G8",
            "G9:G11",
            "G12:G13",
            "G14:G15",
            "G17:K17",
            "G16:I16",
            # f"M{row_num}:N{row_num}",
        ]
        for rng in merged_ranges:
            ws.merge_cells(rng)

        # Base grid border/alignment for the three visible blocks.
        for row in range(1, 19):
            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                c.border = self.border
                if col == 5:
                    c.alignment = self.right
                else:
                    c.alignment = self.center
                c.font = self.normal_font
                if col == 3 or col == 5:
                    c.number_format = '¥#,##0.00'

        for row in range(1, 18):
            for col in range(7, 12):
                c = ws.cell(row=row, column=col)
                c.border = self.border
                if col == 11:
                    c.alignment = self.right
                else:
                    c.alignment = self.center
                c.font = self.normal_font
                if col == 9:
                    c.number_format = '$#,##0.00'
                if col == 11:
                    c.number_format = '¥#,##0.00'

        for row in range(1, row_num + 1):
            for col in range(13, 24):
                c = ws.cell(row=row, column=col)
                c.border = self.border
                c.alignment = self.center
                c.font = self.normal_font
        ws["G17"].number_format = '"运输汇率："0.0000'

        # Section headers and row-2 headers use dark gray fill.
        header_cells = [
            "A1", "G1", "M1", "A2", "B2", "D2", "E2", "G2", "H2", "J2", "K2", "M2", "N2", "O2", "P2", "Q2",
            "R2", "S2", "T2", "U2", "V2", "w2",
        ]
        for cell_ref in header_cells:
            # ws[cell_ref].fill = self.header_fill
            ws[cell_ref].font = self.header_font

        entries = {
            "A1": "国内运费",
            "G1": "国外运费",
            "M1": "运输尺寸",
            "A2": "项目",
            "B2": "单价",
            "D2": "数量",
            "E2": "总金额",
            "G2": "项目",
            "H2": "单价",
            "J2": "数量",
            "K2": "总金额",
            "M2": "序号",
            "N2": "品名",
            "O2": "数量",
            "P2": "件数",
            "Q2": "长(m)",
            "R2": "宽(m)",
            "S2": "高(m)",
            "T2": "总体积（m³）",
            "U2": "单重（kg）",
            "V2": "总重（kg）",
            "W2": "备注",
            "A3": "装箱费",
            "B3": "20GP",
            "C3": 0,
            "D3": "=J3",
            "E3": "=C3*D3",
            "B4": "40GP/HQ",
            "C4": 0,
            "D4": "=J4",
            "E4": "=C4*D4",
            "B5": "40FR",
            "C5": 0,
            "D5": "=J5",
            "E5": "=C5*D5",
            "A6": "港杂费",
            "B6": "20GP",
            "C6": 0,
            "D6": "=J3",
            "E6": "=C6*D6",
            "B7": "40GP/HQ",
            "C7": 0,
            "D7": "=J4",
            "E7": "=C7*D7",
            "B8": "40FR",
            "C8": 0,
            "D8": "=J5",
            "E8": "=C8*D8",
            "A9": "运抵报告费",
            "B9": "20GP",
            "C9": 0,
            "D9": "=J3",
            "E9": "=C9*D9",
            "B10": "40GP/HQ",
            "C10": 0,
            "D10": "=J4",
            "E10": "=C10*D10",
            "B11": "40FR",
            "C11": 0,
            "D11": "=J5",
            "E11": "=C11*D11",
            "A12": "加固费",
            "B12": "每箱",
            "C12": 600,
            "D12": 0,
            "E12": "=C12*D12",
            "A13": "舱单费",
            "B13": "每票",
            "C13": 100,
            "D13": 0,
            "E13": "=C13*D13",
            "A14": "文件费",
            "B14": "每票",
            "C14": 500,
            "D14": 0,
            "E14": "=C14*D14",
            "A15": "报关费",
            "B15": "每票",
            "C15": 300,
            "D15": 0,
            "E15": "=C15*D15",
            "A16": "其他",
            "C16": 0,
            "D16": 0,
            "E16": "=C16*D16",
            "C17": 0,
            "D17": 0,
            "E17": "=C17*D17",
            "A18": "合计",
            "E18": "=SUM(E3:E17)",
            "G3": "海运费",
            "H3": "20GP",
            "I3": 0,
            "J3": 0,
            "K3": "=I3*J3*G17",
            "H4": "40GP/HQ",
            "I4": 0,
            "J4": 0,
            "K4": "=I4*J4*G17",
            "H5": "40FR",
            "I5": 0,
            "J5": 0,
            "K5": "=I5*J5*G17",
            "G6": "DTHC",
            "H6": "20GP",
            "I6": 0,
            "J6": "=J3",
            "K6": "=I6*J6*G17",
            "H7": "40GP/HQ",
            "I7": 0,
            "J7": "=J4",
            "K7": "=I7*J7*G17",
            "H8": "40FR",
            "I8": 0,
            "J8": "=J5",
            "K8": "=I8*J8*G17",
            "G9": "境外陆运",
            "H9": "20GP",
            "I9": 0,
            "J9": "=J3",
            "K9": "=I9*J9*G17",
            "H10": "40GP/HQ",
            "I10": 0,
            "J10": "=J4",
            "K10": "=I10*J10*G17",
            "H11": "40FR",
            "I11": 0,
            "J11": "=J5",
            "K11": "=I11*J11*G17",
            "G12": "电子跟踪单",
            "H12": "5个以内",
            "I12": 0,
            "J12": 0,
            "K12": "=I12*J12*G17",
            "H13": "超过5个",
            "I13": 0,
            "J13": 0,
            "K13": "=I13*J13*G17",
            "G14": "其他",
            "I14": 0,
            "J14": 0,
            "K14": "=I14*J14*G17",
            "I15": 0,
            "J15": 0,
            "K15": "=I15*J15*G17",
            "G16": "合计",
            "K16": "=SUM(K3:K15)",
            "G17": 7,
            f"M{row_num}": "合计",
            f"T{row_num}": f"=SUM(T3:T{row_num - 1})",
            f"V{row_num}": f"=SUM(V3:V{row_num - 1})",
        }
        for cell_ref, value in entries.items():
            ws[cell_ref] = value

        # Populate M/N/O from project commodities (rows 3-14).


        keys = sorted(items.keys())
        for row in range(3, row_num):
            idx = row - 3
            if idx < len(keys):
                key = keys[idx]
                ws[f"M{row}"] = items[key][-1]
                ws[f"N{row}"] = items[key][0]
                ws[f"O{row}"] = self._parse_quantity(items[key][2])
                ws[f"T{row}"] = f"=PRODUCT(Q{row}:S{row})"
                ws[f"V{row}"] = f"=U{row}*P{row}"


    def _build_other_fees(self, ws) -> None:
        ws.title = "其他费用"
        self._set_columns(
            ws,
            {
                "A": 25,
                "B": 24,
                "E": 28,
                "F": 28,
                "G": 26,
                "H": 24,
            },
        )
        for row in range(1, 15):
            ws.row_dimensions[row].height = 24.0
        ws.row_dimensions[15].height = 24.75
        ws.row_dimensions[16].height = 24.75

        for rng in ("A1:B1", "A4:B4", "A8:B8", "A12:B12"):
            ws.merge_cells(rng)

        bold_11 = Font(name=self.normal_font.name, size=11, bold=True)
        bold_12 = Font(name=self.normal_font.name, size=12, bold=True)
        right = Alignment(horizontal="right", vertical="center")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="000000")
        yellow_fill = PatternFill("solid", fgColor="FFFFFF00")

        yellow_amount = {"B2", "B9", "B10"}
        percent_yellow = {"B13", "B6"}
        right_money = {"B2", "B3", "B5", "B6", "B7", "B9", "B10", "B11", "B14"}

        for row in range(2, 15):
            ws[f"A{row}"].border = self.border
            ws[f"A{row}"].alignment = center
            ws[f"A{row}"].font = bold_12
            ws[f"B{row}"].border = self.border
            ws[f"B{row}"].alignment = right if f"B{row}" in right_money else center
            ws[f"B{row}"].font = bold_12 if row in (2, 14) else self.normal_font
            if f"B{row}" in yellow_amount or f"B{row}" in percent_yellow:
                ws[f"B{row}"].fill = yellow_fill
            if f"B{row}" in right_money:
                ws[f"B{row}"].number_format = "#,##0.00"
            if f"B{row}" in percent_yellow:
                ws[f"B{row}"].number_format = "0.00%"

        for row in range(3, 17):
            for col in ("D", "E", "F", "G", "H"):
                cell = ws[f"{col}{row}"]
                cell.border = self.border
                cell.font = self.normal_font
                cell.alignment = center if col == "D" else right
                if col in ("E", "F", "G", "H") and row >= 4:
                    cell.number_format = "#,##0.00"
                if col == "D":
                    cell.font = bold_12

        ws["D16"].font = bold_12
        ws["E3"].font = bold_12
        ws["F3"].font = bold_12
        ws["G3"].font = bold_12
        ws["H3"].font = bold_12

        ws["A1"].border = self.border
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = bold_11
        ws["A4"].border = self.border
        ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A4"].font = bold_11
        ws["B1"].border = Border(top=thin, right=thin, bottom=thin)
        ws["B4"].border = Border(top=thin, right=thin, bottom=thin)
        ws["B8"].border = Border(top=thin, right=thin, bottom=thin)
        ws["B12"].border = Border(top=thin, right=thin, bottom=thin)

        ws["E1"].font = bold_11
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E2"].font = bold_11
        ws["E2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E2"].border = Border(bottom=Side(style="thin", color="000000"))
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F1"].number_format = "#,##0.00"
        ws["F2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H16"].font = bold_11

        entries = {
            "A1": "商检费",
            "E1": "合同总额",
            "A2": "金额",
            "B2": 10000,
            "E2": "付款总额",
            "D3": "月份",
            "E3": "付款",
            "F3": "收款",
            "G3": "结余",
            "H3": "占用费",
            "A4": "保险费",
            "D4": 1,
            "E4": 0,
            "F4": 0,
            "G4": "=E4-F4",
            "H4": "=IF(G4>0,G4*B$13/12,0)",
            "A5": "对外金额",
            "B5": self.project.totalsum,
            "D5": 2,
            "E5": 0,
            "F5": 0,
            "G5": "=E5-F5+G4",
            "H5": "=IF(G5>0,G5*B$13/12,0)",
            "A6": "费率",
            "B6": 0.001,
            "D6": 3,
            "E6": 0,
            "F6": 0,
            "G6": "=E6-F6+G5",
            "H6": "=IF(G6>0,G6*B$13/12,0)",
            "A7": "保险费",
            "B7": "=B5*1.1*B6",
            "D7": 4,
            "E7": 0,
            "F7": 0,
            "G7": "=E7-F7+G6",
            "H7": "=IF(G7>0,G7*B$13/12,0)",
            "A8": "其他费用",
            "D8": 5,
            "E8": 0,
            "F8": 0,
            "G8": "=E8-F8+G7",
            "H8": "=IF(G8>0,G8*B$13/12,0)",
            "A9": "须中方承担的其他费用",
            "B9": 0,
            "D9": 6,
            "E9": 0,
            "F9": 0,
            "G9": "=E9-F9+G8",
            "H9": "=IF(G9>0,G9*B$13/12,0)",
            "A10": "安放费",
            "B10": 0,
            "D10": 7,
            "E10": 0,
            "F10": 0,
            "G10": "=E10-F10+G9",
            "H10": "=IF(G10>0,G10*B$13/12,0)",
            "A11": "合计",
            "B11": "=SUM(B9:B10)",
            "D11": 8,
            "E11": 0,
            "F11": 0,
            "G11": "=E11-F11+G10",
            "H11": "=IF(G11>0,G11*B$13/12,0)",
            "A12": "资金占用费",
            "D12": 9,
            "E12": 0,
            "F12": 0,
            "G12": "=E12-F12+G11",
            "H12": "=IF(G12>0,G12*B$13/12,0)",
            "A13": "年化利率",
            "B13": 0.025,
            "D13": 10,
            "E13": 0,
            "F13": 0,
            "G13": "=E13-F13+G12",
            "H13": "=IF(G13>0,G13*B$13/12,0)",
            "A14": "资金占用费总额",
            "B14": "=round(H16,0)",
            "D14": 11,
            "E14": 0,
            "F14": 0,
            "G14": "=E14-F14+G13",
            "H14": "=IF(G14>0,G14*B$13/12,0)",
            "D15": 12,
            "E15": 0,
            "F15": 0,
            "G15": "=E15-F15+G14",
            "H15": "=IF(G15>0,G15*B$13/12,0)",
            "D16": "合计",
            "H16": "=SUM(H4:H15)",
        }
        for cell_ref, value in entries.items():
            ws[cell_ref] = value

        ws["B14"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws["B2"].font = bold_11
        ws["B13"].alignment = Alignment(vertical="center")
        ws["E3"].alignment = center
        ws["F3"].alignment = center
        ws["G3"].alignment = center
        ws["H3"].alignment = center
        ws["E3"].border = Border(
            left=thin,
            right=thin,
            bottom=thin,
        )
        for ref in ("E16", "F16", "G16"):
            ws[ref].number_format = "General"
            ws[ref].alignment = Alignment()


    def _build_inner_quote(self, ws, item_count: int) -> int:
        ws.title = "2.物资对内分项报价表"
        self._set_columns(
            ws,
            {
                "A": 11.25,
                "B": 29.75,
                "C": 14.875,
                "D": 11.75,
                "E": 17.625,
                "F": 15.0,
                "G": 10.625,
                "H": 10.125,
                "I": 16.625,
                "J": 16.0,
                "K": 15.0,
                "L": 16.75,
                "M": 15.75,
                "N": 19.625,
            },
        )

        item_start = 5
        item_end = item_start + item_count - 1
        total_row = item_end + 1
        stamp_row = total_row + 2
        date_row = total_row + 3
        summary_row = total_row + 5
        service_row = total_row + 6

        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 24
        for row in range(4, total_row + 1):
            ws.row_dimensions[row].height = 24.0
        ws.row_dimensions[total_row + 1].height = 32
        ws.row_dimensions[stamp_row].height = 33
        ws.row_dimensions[date_row].height = 33
        ws.row_dimensions[summary_row - 1].height = 20
        ws.row_dimensions[summary_row].height = 20
        ws.row_dimensions[service_row].height = 15

        ws.merge_cells("A1:N1")
        ws.merge_cells("A3:B4")
        ws.merge_cells("C3:E3")
        for col in "FGHIJKLMN":
            ws.merge_cells(f"{col}3:{col}4")
        ws.merge_cells(f"A{item_start}:A{item_end}")
        ws.merge_cells(f"A{total_row}:D{total_row}")


        ws["A1"] = "二.物资对内分项报价表"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center

        ws.merge_cells("A2:N2")
        ws["A2"] = "报价单位：人民币元（保留小数点后两位）"
        ws["A2"].font = self.normal_font
        ws["A2"].alignment = self.left

        ws["A3"] = "物资"
        ws["C3"] = "商品购买价款"
        ws["C4"] = "单价"
        ws["D4"] = "数量"
        ws["E4"] = "总价"
        ws["F3"] = "国内运杂费"
        ws["G3"] = "包装费"
        ws["H3"] = "保管费"
        ws["I3"] = "物资检验费"
        ws["J3"] = "运输保险费"
        ws["K3"] = "国外运费"
        ws["L3"] = "实施服务费"
        ws["M3"] = "税金"
        ws["N3"] = "合计"

        self._style_row(ws, 3, 1, 14, header=True)
        self._style_row(ws, 4, 1, 14, header=True)

        ws[f"A{item_start}"] = "供货清单（一）"
        ws[f"A{item_start}"].alignment = self.center
        ws[f"A{item_start}"].font = self.normal_font
        ws[f"A{item_start}"].border = self.border

        for i in range(1, item_count + 1):
            row = item_start + i - 1
            ws[f"B{row}"] = f'=物资选择!A{i}&"."&物资选择!B{i}'
            ws[f"C{row}"] = f"=INDEX(全部厂家备用!J1:J{item_count + 5},物资选择!C{i})"
            ws[f"D{row}"] = f"=INDEX(全部厂家备用!E1:E{item_count + 5},物资选择!C{i})"
            ws[f"E{row}"] = f"=C{row}*D{row}"
            ws[f"F{row}"] = 0
            ws[f"G{row}"] = 0
            ws[f"H{row}"] = 0
            ws[f"I{row}"] = f"=ROUND(E{row}/E${total_row}*I${summary_row},2)"
            ws[f"J{row}"] = f"=ROUND(E{row}/E${total_row}*J${summary_row},2)"
            ws[f"K{row}"] = f"=ROUND(E{row}/E${total_row}*K${summary_row},2)"
            ws[f"L{row}"] = f"=ROUND(E{row}/E${total_row}*L${summary_row},2)"
            ws[f"M{row}"] = f"=ROUND(E{row}/E${total_row}*M${summary_row},2)"
            ws[f"N{row}"] = f"=SUM(E{row}:M{row})"
            self._style_row(ws, row, 2, 14, header=False)
            ws[f"B{row}"].alignment = self.left
            ws[f"C{row}"].number_format = "#,##0.00"
            ws[f"D{row}"].number_format = "0"
            for col in "EFGHIJKLMN":
                ws[f"{col}{row}"].number_format = "#,##0.00"
            for col in "CEFGHIJKLMN":
                ws[f"{col}{row}"].alignment = self.right

        ws[f"A{total_row}"] = "合计"
        ws[f"A{total_row}"].font = self.header_font
        ws[f"A{total_row}"].alignment = self.center
        for col in "ABCD":
            ws[f"{col}{total_row}"].border = self.border
        for col in "EFGHIJKLMN":
            ws[f"{col}{total_row}"] = f"=SUM({col}{item_start}:{col}{item_end})"
            ws[f"{col}{total_row}"].number_format = "#,##0.00"
            ws[f"{col}{total_row}"].alignment = self.right
            ws[f"{col}{total_row}"].border = self.border
            ws[f"{col}{total_row}"].font = self.header_font


        ws[f"L{stamp_row}"] = "投标人盖章："
        ws[f"L{stamp_row}"].font = self.header_font
        ws[f"L{stamp_row}"].alignment = self.right
        ws[f"L{date_row}"] = "日期："
        ws[f"L{date_row}"].font = self.header_font
        ws[f"L{date_row}"].alignment = self.right
        ws[f"M{date_row}"].number_format = 'yyyy"年"m"月"d"日"'
        ws[f"M{date_row}"] = self._parse_date(self.project.date)
        ws[f"M{date_row}"].font = self.header_font
        ws[f"M{date_row}"].alignment = self.right

        ws[f"F{summary_row}"] = 0
        ws[f"I{summary_row}"] = "=其他费用!B2"
        ws[f"J{summary_row}"] = "=其他费用!B7"
        ws[f"K{summary_row}"] = "=运输费用!K16+运输费用!E18"
        ws[f"L{summary_row}"] = f"=L{service_row}"
        ws[f"M{summary_row}"] = f"=ROUND((SUM(E{total_row}:K{total_row}))*0.0003,2)"
        ws[f"N{summary_row}"] = f"=SUM(E{total_row}:M{total_row})"
        ws[f"L{service_row}"] = (
            f"=IF(E{total_row}>50000000,(E{total_row}-50000000)*0.0075+835000,"
            f"IF(E{total_row}>20000000,(E{total_row}-20000000)*0.01+535000,"
            f"IF(E{total_row}>10000000,(E{total_row}-10000000)*0.02+335000,"
            f"IF(E{total_row}>5000000,(E{total_row}-5000000)*0.03+185000,"
            f"IF(E{total_row}>2000000,(E{total_row}-2000000)*0.035+80000,E{total_row}*0.04)))))"
        )

        # self._style_row(ws, summary_row, 6, 14, header=False)
        # ws[f"F{summary_row}"].number_format = "#,##0.00"
        # for col in "IJKLMN":
        #     ws[f"{col}{summary_row}"].number_format = "#,##0.00"
        # ws[f"L{service_row}"].number_format = "#,##0.00"
        # ws[f"L{service_row}"].border = self.border

        return total_row


    def _build_tax_sheet(self, ws, item_count: int, inner_total_row: int) -> None:
        ws.title = "3.各项物资退抵税额表"
        self._set_columns(
            ws,
            {
                "A": 8.0,
                "B": 28,
                "C": 28,
                "D": 15,
                "E": 16,
                "F": 15,
                "G": 20,
                "H": 22,
            },
        )

        item_start = 4
        item_end = item_start + item_count - 1
        trans_row = item_end + 1
        insurance_row = item_end + 2
        inspect_row = item_end + 3
        total_row = item_end + 4
        promise_row = total_row + 1
        spacer_row = total_row + 2
        stamp_row = total_row + 3
        date_row = total_row + 4

        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 75
        for r in range(item_start, total_row + 1):
            ws.row_dimensions[r].height = 25
        ws.row_dimensions[promise_row].height = 57
        ws.row_dimensions[spacer_row].height = 22
        ws.row_dimensions[stamp_row].height = 30
        ws.row_dimensions[date_row].height = 35

        ws.merge_cells("A1:H1")
        ws.merge_cells("A2:C2")
        ws.merge_cells(f"A{promise_row}:H{promise_row}")

        ws["A1"] = "三.增值税和消费税退抵税额表"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center

        ws["A2"] = "报价单位：人民币元（保留小数点后两位）"
        ws["A2"].font = self.normal_font
        ws["A2"].alignment = self.left

        headers = [
            "序号",
            "品名",
            "投标人向物资生产供货企业、运输企业、第三方检验机构、土建涉及施工单位支付的含税商品或服务购买价款",
            "投标人预期可获得的退抵增值税率(%)",
            "投标人预期可获得的退抵增值税额",
            "投标人预期可获得的退抵消费税率(%)",
            "投标人预期可获得的退抵消费税额",
            "投标人预期可获得的退抵增值税和消费税总额",
        ]
        for c, h in enumerate(headers, start=1):
            ws.cell(3, c, h)
        self._style_row(ws, 3, 1, 8, header=True)

        for i in range(1, item_count + 1):
            row = item_start + i - 1
            ws[f"A{row}"] = f"=物资选择!A{i}"
            ws[f"B{row}"] = f"=物资选择!B{i}"
            ws[f"C{row}"] = f"='2.物资对内分项报价表'!E{4 + i}"
            ws[f"D{row}"] = 13
            ws[f"E{row}"] = f"=ROUND(C{row}/(1+D{row}/100)*D{row}/100,2)"
            ws[f"F{row}"] = 0
            ws[f"G{row}"] = f"=ROUND(C{row}/(1+F{row}/100)*F{row}/100,2)"
            ws[f"H{row}"] = f"=E{row}+G{row}"
            self._style_row(ws, row, 1, 8, header=False)

        ws[f"B{trans_row}"] = "运输"
        ws[f"C{trans_row}"] = f"='2.物资对内分项报价表'!K{inner_total_row}"
        ws[f"D{trans_row}"] = 0
        ws[f"E{trans_row}"] = f"=ROUND(C{trans_row}/(1+D{trans_row}/100)*D{trans_row}/100,2)"
        ws[f"F{trans_row}"] = 0
        ws[f"G{trans_row}"] = f"=ROUND(C{trans_row}/(1+F{trans_row}/100)*F{trans_row}/100,2)"
        ws[f"H{trans_row}"] = f"=E{trans_row}+G{trans_row}"

        ws[f"B{insurance_row}"] = "保险"
        ws[f"C{insurance_row}"] = f"='2.物资对内分项报价表'!J{inner_total_row}"
        ws[f"D{insurance_row}"] = 0
        ws[f"E{insurance_row}"] = f"=ROUND(C{insurance_row}/(1+D{insurance_row}/100)*D{insurance_row}/100,2)"
        ws[f"F{insurance_row}"] = 0
        ws[f"G{insurance_row}"] = f"=ROUND(C{insurance_row}/(1+F{insurance_row}/100)*F{insurance_row}/100,2)"
        ws[f"H{insurance_row}"] = f"=E{insurance_row}+G{insurance_row}"

        ws[f"B{inspect_row}"] = "第三方检验"
        ws[f"C{inspect_row}"] = f"='2.物资对内分项报价表'!I{inner_total_row}"
        ws[f"D{inspect_row}"] = 0
        ws[f"E{inspect_row}"] = f"=ROUND(C{inspect_row}/(1+D{inspect_row}/100)*D{inspect_row}/100,2)"
        ws[f"F{inspect_row}"] = 0
        ws[f"G{inspect_row}"] = f"=ROUND(C{inspect_row}/(1+F{inspect_row}/100)*F{inspect_row}/100,2)"
        ws[f"H{inspect_row}"] = f"=E{inspect_row}+G{inspect_row}"

        for row in range(4, total_row + 1):
            for col in "CEGH":
                ws[f"{col}{row}"].number_format = "#,##0.00"

        self._style_row(ws, trans_row, 1, 8, header=False)
        self._style_row(ws, insurance_row, 1, 8, header=False)
        self._style_row(ws, inspect_row, 1, 8, header=False)

        ws[f"A{total_row}"] = "共计"
        ws.merge_cells(f"A{total_row}:B{total_row}")
        ws[f"A{total_row}"].font = self.header_font
        ws[f"A{total_row}"].alignment = self.center
        ws[f"A{total_row}"].border = self.border
        ws[f"B{total_row}"].border = self.border
        ws[f"C{total_row}"] = f"=SUM(C{item_start}:C{inspect_row})"
        ws[f"E{total_row}"] = f"=SUM(E{item_start}:E{inspect_row})"
        ws[f"G{total_row}"] = f"=SUM(G{item_start}:G{inspect_row})"
        ws[f"H{total_row}"] = f"=SUM(H{item_start}:H{inspect_row})"
        self._style_row(ws, total_row, 3, 8, header=True)

        ws[f"A{promise_row}"] = "我公司承诺：如我公司中标, 我公司出现与投标文件承诺的增值税或消费税退税或抵扣内容不一致的情况，实际退税或抵扣金额高于投标文件中承诺金额的，我公司会将超出部分及其孳息退回采购人指定账户，并接受违规、违约处理。"
        ws[f"A{promise_row}"].font = self.header_font
        ws[f"A{promise_row}"].alignment = self.left
        ws[f"A{promise_row}"].border = self.border
        for c in range(2, 9):
            ws.cell(promise_row, c).border = self.border

        ws[f"G{stamp_row}"] = "投标人盖章："
        ws[f"G{stamp_row}"].font = self.header_font
        ws[f"G{stamp_row}"].alignment = self.right
        ws[f"G{date_row}"] = "日期："
        ws[f"G{date_row}"].font = self.header_font
        ws[f"G{date_row}"].alignment = self.right
        ws[f"H{date_row}"] = self._parse_date(self.project.date)
        ws[f"H{date_row}"].font = self.header_font
        ws[f"H{date_row}"].alignment = self.right
        ws[f"H{date_row}"].number_format = 'yyyy"年"m"月"d"日"'

        self._tax_total_row = total_row


    def _build_tech_sheet(self, ws) -> None:
        if not self.project.is_tech:
            return

        people = self.project.techinfo[0] if len(self.project.techinfo) >= 2 else 0
        days = self.project.techinfo[1] if len(self.project.techinfo) >= 2 else 0

        ws.title = "4.技术服务费报价表"
        self._set_columns(
            ws,
            {
                "A": 6.0,
                "B": 16.0,
                "C": 14.0,
                "D": 16.0,
                "E": 8.0,
                "G": 18,
                "H": 22,
                "I": 9.0,
            },
        )

        ws.row_dimensions[1].height = 52
        for row in range(2, 15):
            ws.row_dimensions[row].height = 42
        ws.row_dimensions[17].height = 24
        ws.row_dimensions[18].height = 24

        ws["A1"] = "四.技术服务费报价表"
        ws["A1"].font = Font(name="宋体", size=20, bold=True)
        ws["A1"].alignment = self.center

        headers = ["序号", "费用名称", "美元单价", "人民币单价", "人数", "天/次数", "美元合计", "人民币合计"]
        for c, h in enumerate(headers, start=1):
            ws.cell(2, c, h)
        self._style_row(ws, 2, 1, 8, header=True)
        # self._style_row(ws, 2, 7, 8, header=True)
        # ws["G2"].alignment = Alignment(horizontal="center", vertical="center")
        # ws["H2"].alignment = Alignment(horizontal="center", vertical="center")

        rows = [
            (1, "护照和签证手续费", 0, "-", 0, "=D3*E3"),
            (2, "防疫免疫费", 0, "-", 0, "=D4*E4"),
            (3, "技术服务人员保险费", 800, "-", 0, "=D5*E5"),
            (4, "国内交通费", 0, "-", 0, "=D6*E6"),
            (5, "国际交通费", 10000, "-", 0, "=D7*E7"),
            (6, "住宿费", 500, days, "=C8*E8*F8", "=D8*E8*F8"),
            (7, "伙食费", 50, 15, "=C9*E9*F9", "=D9*E9*F9"),
            (8, "津贴补贴", 50, 15, "=C10*E10*F10", "=D10*E10*F10"),
            (9, "当地雇工费", None, None, 0, 0),
            (10, "当地设备工具材料购置或租用费", None, None, 0, 0),
            (11, "其它确需发生的费用", None, None, 0, 0),
        ]

        for idx, row in enumerate(rows, start=3):
            seq, name, rmb_price, freq, usd_total, rmb_total = row
            ws[f"A{idx}"] = seq
            ws[f"B{idx}"] = name
            ws[f"G{idx}"] = usd_total
            ws[f"H{idx}"] = rmb_total

            self._style_row(ws, idx, 1, 8, header=False)
            ws[f"B{idx}"].alignment = self.left
            ws[f"G{idx}"].number_format = '$#,##0.00'
            ws[f"H{idx}"].number_format = '¥#,##0.00'

            if idx <= 10:
                ws[f"D{idx}"] = rmb_price
                ws[f"E{idx}"] = people
                ws[f"F{idx}"] = freq
                ws[f"D{idx}"].number_format = '¥#,##0.00'
            if idx <= 7:
                ws[f"F{idx}"] = "-"
            if idx in (8, 9, 10):
                ws[f"F{idx}"] = days
                ws[f"F{idx}"].number_format = "0"
            if idx in (11, 12, 13):
                for col in range(3, 7):
                    ws.cell(idx, col).border = self.border

        for row in range(3, 14):
            ws[f"A{row}"].number_format = "0"
            ws[f"E{row}"].number_format = "0"
            ws[f"F{row}"].alignment = self.center
            ws[f"H{row}"].alignment = self.right
            ws[f"C{row}"].border = self.diag_border

        for row in range(11, 14):
            for col in 'DEF':
                ws[f"{col}{row}"].border = self.diag_border

        ws.merge_cells("A1:H1")
        ws.merge_cells("C11:F11")
        ws.merge_cells("C12:F12")
        ws.merge_cells("C13:F13")
        ws.merge_cells("A14:F14")

        ws["A14"] = "共计："
        ws["A14"].font = self.header_font
        ws["A14"].alignment = self.center
        for col in range(1, 7):
            ws.cell(14, col).border = self.border
            ws.cell(14, col).font = self.header_font
        ws["G14"] = "=SUM(G3:G13)"
        ws["H14"] = "=SUM(H3:H13)"
        ws["G14"].number_format = '$#,##0.00'
        ws["H14"].number_format = '¥#,##0.00'
        ws["G14"].alignment = self.center
        ws["H14"].alignment = self.right
        ws["G14"].border = self.border
        ws["H14"].border = self.border
        ws["G14"].font = self.header_font
        ws["H14"].font = self.header_font

        ws["G17"] = "投标人盖章："
        ws["G17"].font = self.header_font
        ws["G17"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws["G18"] = "日期："
        ws["G18"].font = self.header_font
        ws["G18"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws["H18"] = self._parse_date(self.project.date)
        ws["H18"].font = self.header_font
        ws["H18"].alignment = self.right
        ws["H18"].number_format = 'yyyy"年"m"月"d"日"'

    def _build_train_sheet(self, ws) -> None:
        if not self.project.is_cc:
            return

        title_no = "5" if self.project.is_tech else "4"
        num = self.project.training_num
        days = self.project.training_days
        stay_days = max(days - 1, 0)

        ws.title = f"{title_no}.来华培训费报价表"
        self._set_columns(
            ws,
            {
                "A": 6.0,
                "B": 15.0,
                "C": 8.0,
                "D": 16,
                "E": 7.0,
                "F": 16,
                "G": 20,
                "H": 12.0,
                "I": 9.0,
            },
        )

        ws.row_dimensions[1].height = 30.0
        for row in range(2, 18):
            ws.row_dimensions[row].height = 32
        ws.row_dimensions[18].height = 20
        ws.row_dimensions[19].height = 20
        ws.row_dimensions[20].height = 20

        for rng in (
            "A1:H1",
            "A2:A3",
            "B2:C3",
            "D2:F2",
            "G2:G3",
            "H2:H3",
            "B4:C4",
            "B5:C5",
            "D5:G5",
            "B6:C6",
            "B7:C7",
            "B8:C8",
            "B9:C9",
            "B10:C10",
            "B11:C11",
            "B12:C12",
            "B13:C13",
            "D13:G13",
            "B14:C14",
            "D14:F14",
            "A15:A16",
            "B15:B16",
            "B17:F17",
        ):
            ws.merge_cells(rng)
        title_char = "五" if self.project.is_tech else "四"
        ws["A1"] = f"{title_char}.来华培训费报价表"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center

        headers = {
            "A2": "序号",
            "B2": "费用名称",
            "D2": "费用计算方式",
            "G2": "人民币（元）",
            "H2": "其中含购汇人民币限额",
            "D3": "标准",
            "E3": "人数",
            "F3": "天（次）数",
        }
        for cell_ref, value in headers.items():
            ws[cell_ref] = value
        self._style_row(ws, 2, 1, 8, header=True)
        self._style_row(ws, 3, 1, 8, header=True)


        rows = {
            4: ("一", "培训费", None, 360, num, days, "=D4*E4*F4"),
            5: ("二", "接待费", None, None, None, None, None),
            6: (1, "日常伙食费", None, 190, num, days, "=D6*E6*F6"),
            7: (2, "住宿费", None, 350, num, stay_days, "=D7*E7*F7"),
            8: (3, "宴请费", None, 200, num, 1, "=D8*E8*F8"),
            9: (4, "零用费", None, 150, num, days, "=D9*E9*F9"),
            10: (5, "小礼品费", None, 200, num, 1, "=D10*E10*F10"),
            11: (6, "人身意外伤害保险", None, 150, num, "-", "=D11*E11"),
            12: ("三", "国际旅费", None, 5000, num, "-", "=D12*E12"),
            13: ("四", "管理费", None, None, None, None, None),
            14: (1, "承办管理费", None, None, None, None, "=ROUND((SUM(G4,G6:G11))*0.06,2)"),
            15: (2, "管理人员费", "伙食费", "=D6", 2, days, "=D15*E15*F15"),
            16: (None, None, "住宿费", "=D7", 2, days, "=D16*E16*F16"),
            17: ("五", "合计", None, None, None, None, "=SUM(G4:G16)"),
        }

        for row, values in rows.items():
            seq, name_b, name_c, standard, people, freq, total = values
            if seq is not None:
                ws[f"A{row}"] = seq
            if name_b is not None:
                ws[f"B{row}"] = name_b
            if name_c is not None:
                ws[f"C{row}"] = name_c
            if standard is not None:
                ws[f"D{row}"] = standard
            if people is not None:
                ws[f"E{row}"] = people
            if freq is not None:
                ws[f"F{row}"] = freq
            if total is not None:
                ws[f"G{row}"] = total

            self._style_row(ws, row, 1, 8, header=False)
            for col in range(1, 9):
                ws.cell(row, col).border = self.border

        for ref in ("B4", "B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14", "B15","C15", "C16"):
            ws[ref].alignment = self.left

        for ref in ("G4", "G6", "G7", "G8", "G9", "G10", "G11", "G12", "G14", "G15", "G16", "G17"):
            ws[ref].number_format = '¥#,##0.00'
            ws[ref].alignment = self.right

        for ref in ("E4", "E6", "E7", "E8", "E9", "E10", "E11", "E12", "E15", "E16", "F4", "F6", "F7", "F8", "F9", "F10", "F15", "F16"):
            ws[ref].number_format = "0"
            ws[ref].alignment = self.center

        ws["D4"].number_format = '0"元/人*天"'
        ws["D6"].number_format = '0"元/人*天"'
        ws["D7"].number_format = '0"元/人*天"'
        ws["D8"].number_format = '0"元/人*次"'
        ws["D9"].number_format = '0"元/人*天"'
        ws["D10"].number_format = '0"元/人"'
        ws["D11"].number_format = '0"元/人"'
        ws["D12"].number_format = '0"元/人"'
        ws["D15"].number_format = '0"元/人*天"'
        ws["D16"].number_format = '0"元/人*天"'

        for col in 'DEF':
            ws[f"{col}14"].border = self.diag_border

        self._style_row(ws, 17, 1, 8, header=True)

        ws["F19"] = "投标人盖章："
        ws["F19"].font = self.header_font
        ws["F19"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws["F20"] = "日期："
        ws["F20"].font = self.header_font
        ws["F20"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws["G20"] = self._parse_date(self.project.date)
        ws["G20"].font = self.header_font
        ws["G20"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws["G20"].number_format = 'yyyy"年"m"月"d"日"'


    def _build_total_sheet(
        self, ws, inner_total_row: int, tax_total_row: int, bid_date: date
    ) -> None:
        ws.title = "1.投标报价总表"
        self._set_columns(ws, {"A": 10.0, "B": 35.0, "C": 24, "D": 20})

        ws.row_dimensions[1].height = 50
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 40

        rows = [
            (
                "一",
                "全部物资价格（含商品购买价款、国内运杂费、包装费、保管费、物资检验费、运输保险费、国外运费、实施服务费、税金）",
                f"='2.物资对内分项报价表'!N{inner_total_row}",
                f"{self.project.trans}{self.project.destination}",
            )
        ]
        seq = ["一", "二", "三", "四", "五"]
        if self.project.is_tech:
            rows.append((seq[len(rows)], "技术服务费", "='4.技术服务费报价表'!H14", ""))
        if self.project.is_cc:
            train_sheet = "5.来华培训费报价表" if self.project.is_tech else "4.来华培训费报价表"
            rows.append((seq[len(rows)], "来华培训和接待费", f"='{train_sheet}'!G17", ""))
        rows.append((seq[len(rows)], "其他费用-安防费用", "=其他费用!B11", ""))
        rows.append(
            (
                seq[len(rows)],
                "《供货清单（一）》中各项物资以及全部物资运输、检验、配套土建（本项目不适用）等的购买价款中的增值税和消费税退抵税总额",
                f"='3.各项物资退抵税额表'!H{tax_total_row}",
                "",
            )
        )

        detail_start = 4
        detail_heights = [108, 108, 64, 64, 122]
        for idx, (no, item, amount, note) in enumerate(rows, start=detail_start):
            ws.row_dimensions[idx].height = detail_heights[idx - detail_start]
            ws[f"A{idx}"] = no
            ws[f"B{idx}"] = item
            ws[f"C{idx}"] = amount
            ws[f"D{idx}"] = note
            self._style_row(ws, idx, 1, 4, header=False)
            ws[f"B{idx}"].alignment = self.left
            ws[f"C{idx}"].alignment = self.right
            ws[f"C{idx}"].number_format = '¥#,##0.00'
            ws[f"D{idx}"].alignment = self.left

        total_row = detail_start + len(rows)
        self._total_sheet_total_row = total_row
        note_row = total_row + 1
        option1_row = total_row + 2
        option2_row = total_row + 3
        stamp_row = total_row + 4
        date_row = total_row + 5

        ws.row_dimensions[total_row].height = 40
        ws.row_dimensions[note_row].height = 70
        ws.row_dimensions[option1_row].height = 50
        ws.row_dimensions[option2_row].height = 50
        ws.row_dimensions[stamp_row].height = 30
        ws.row_dimensions[date_row].height = 30

        ws.merge_cells("A1:D1")
        ws.merge_cells("A2:D2")
        ws.merge_cells(f"A{note_row}:D{note_row}")
        ws.merge_cells(f"B{option1_row}:D{option1_row}")
        ws.merge_cells(f"B{option2_row}:D{option2_row}")

        ws["A1"] = "一.投标报价总表"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center

        ws["A2"] = "报价单位：人民币元（保留小数点后两位）"
        ws["A2"].font = self.normal_font
        ws["A2"].alignment = self.left

        for c, h in enumerate(["序号", "费用项目", "合计金额", "备注"], start=1):
            ws.cell(3, c, h)
        self._style_row(ws, 3, 1, 4, header=True)

        tax_row = detail_start + len(rows) - 1
        positive_end_row = tax_row - 1
        ws[f"B{total_row}"] = "共计"
        ws[f"C{total_row}"] = f"=SUM(C{detail_start}:C{positive_end_row})-C{tax_row}"
        self._style_row(ws, total_row, 2, 3, header=False)
        ws[f"B{total_row}"].font = self.header_font
        ws[f"C{total_row}"].font = self.header_font
        ws[f"B{total_row}"].alignment = self.center
        ws[f"C{total_row}"].alignment = self.right
        ws[f"C{total_row}"].number_format = '¥#,##0.00'
        ws[f"D{total_row}"].border = self.border

        ws[f"A{note_row}"] = "如中标，采购人向我公司支付中标价中的外汇时，汇率按照以下第一种方式确定（投标人未明确汇率确定方式的，则视投标人选择第一种方式；如中标，投标人须在与采购人商签合同期间，书面向采购人提供收款账户开户银行信息）："
        ws[f"A{note_row}"].font = self.header_font
        ws[f"A{note_row}"].alignment = self.left
        for col in range(1, 5):
            ws.cell(note_row, col).border = self.border

        ws[f"A{option1_row}"] = "一"
        ws[f"B{option1_row}"] = "按照结算当日，我公司收款账户开户银行公布的现汇卖出价为准。我公司收款账户开户银行为：招商银行北京分行"
        self._style_row(ws, option1_row, 1, 4, header=False)
        ws[f"B{option1_row}"].alignment = self.left

        ws[f"A{option2_row}"] = "二"
        ws[f"B{option2_row}"] = "以我公司本次投标中使用的汇率，即    ，作为采购人向我公司结算外汇时使用的汇率。"
        self._style_row(ws, option2_row, 1, 4, header=False)
        ws[f"B{option2_row}"].alignment = self.left

        ws[f"C{stamp_row}"] = "投标人盖章："
        ws[f"C{stamp_row}"].font = self.header_font
        ws[f"C{stamp_row}"].alignment = self.right
        ws[f"C{date_row}"] = "日期："
        ws[f"C{date_row}"].font = self.header_font
        ws[f"C{date_row}"].alignment = self.right
        ws[f"D{date_row}"] = bid_date
        ws[f"D{date_row}"].alignment = self.right
        ws[f"D{date_row}"].font = self.header_font
        ws[f"D{date_row}"].number_format = 'yyyy"年"m"月"d"日"'

    def _build_opening_sheet(self, ws, bid_date: date) -> None:
        ws.title = "3.开标一览表"
        self._set_columns(ws, {"A": 19.375, "B": 25.0, "C": 56.25, "D": 25.0})

        ws.row_dimensions[1].height = 50
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 30
        ws.row_dimensions[5].height = 60
        ws.row_dimensions[9].height = 25
        ws.row_dimensions[10].height = 25

        ws.merge_cells("A1:D1")
        ws.merge_cells("C2:D2")
        ws.merge_cells("A3:A4")
        ws.merge_cells("C3:C4")
        ws.merge_cells("D3:D4")

        ws["A1"] = "三、开标一览表"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center

        ws["A2"] = "招标编号："
        ws["B2"] = self.project.code
        ws["C2"] = f"项目名称：{self.project.name}"
        for ref in ("A2", "B2", "C2"):
            ws[ref].font = self.normal_font
            ws[ref].alignment = self.right if ref in ("A2", "C2") else self.left

        ws["A3"] = "投标人名称"
        ws["B3"] = "投标报价"
        ws["C3"] = "启运或运抵时间"
        ws["D3"] = "备注"
        self._style_row(ws, 3, 1, 4, header=True)
        self._style_row(ws, 4, 1, 4, header=True)

        ws["B4"] = "小写￥"
        ws["A5"] = "中国海外经济合作有限公司"
        ws["B5"] = f"='1.投标报价总表'!C{self._total_sheet_total_row}"
        ws["C5"] = self.project.trans_time
        ws["C5"].fill = self.import_fill
        ws["D5"] = ""
        self._style_row(ws, 5, 1, 4, header=False)
        ws["A5"].alignment = self.left
        ws["C5"].alignment = self.left
        ws["B5"].alignment = self.right
        ws["B5"].number_format = '¥#,##0.00'

        ws["C9"] = "投标人盖章："
        ws["C10"] = "日期："
        ws["D10"] = bid_date
        ws["C9"].font = self.header_font
        ws["C10"].font = self.header_font
        ws["D10"].font = self.header_font
        ws["D10"].number_format = 'yyyy"年"m"月"d"日"'
        ws["C9"].alignment = self.right
        ws["C10"].alignment = self.right
        ws["D10"].alignment = self.left

    def _build_procurement_deviation_sheet(self, ws, items: Dict) -> None:
        ws.title = "2.采购需求偏离表(物资部分)"
        supplier_last_row = max(getattr(self, "_all_suppliers_last_row", 1), 1)
        sorted_keys = sorted(items.keys())
        small_font = Font(name="宋体", size=10)
        title_font = Font(name="宋体", size=14, bold=True)
        left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        footer_font = Font(name="宋体", size=14, bold=True)

        self._set_columns(
            ws,
            {
                "A": 7,
                "B": 10,
                "C": 19,
                "D": 60,
                "E": 19,
                "F": 60,
                "G": 10,
                "H": 10,
            },
        )

        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 24
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 28
        ws.row_dimensions[5].height = 28

        merged_ranges = [
            "A1:H1",
            "A2:B2",
            "A3:B3",
            "C3:D3",
            "C4:D4",
            "E4:F4",
            "A5:H5",
        ]

        label_rows = [
            "品名",
            "数量及单位",
            "HS编码",
            "具体规格、参数和功能表述",
            "检验标准",
            "品牌",
            "生产企业",
            "供货企业",
            "强制性认证取证情况",
        ]
        response_labels = [
            "品名",
            "数量及单位",
            "HS编码",
            "产品型号、参数和功能表述",
            "检验标准",
            "品牌",
            "生产企业",
            "供货企业",
            "强制性认证取证情况",
        ]

        for idx, key in enumerate(sorted_keys):
            row_base = 6 + idx * 9
            spec_text = items[key][4] if len(items[key]) > 4 else ""
            spec_lines = max(str(spec_text).count("\n") + 1, 1)
            for offset in range(9):
                ws.row_dimensions[row_base + offset].height = 15
            ws.row_dimensions[row_base + 3].height = max(30, spec_lines * 14)
            merged_ranges.append(f"A{row_base}:A{row_base + 8}")
            merged_ranges.append(f"B{row_base}:B{row_base + 8}")

        blank_row = 6 + len(sorted_keys) * 9
        stamp_row = blank_row + 1
        date_row = blank_row + 2
        ws.row_dimensions[stamp_row].height = 20
        ws.row_dimensions[date_row].height = 20
        merged_ranges.append(f"G{date_row}:H{date_row}")

        for cell_range in merged_ranges:
            ws.merge_cells(cell_range)

        ws["A1"] = "采购需求偏离表（物资部分）"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_wrap

        ws["A2"] = "招标编号："
        ws["C2"] = self.project.code
        ws["A3"] = "项目名称："
        ws["C3"] = self.project.name
        ws["A4"] = "序号"
        ws["B4"] = "采购需求条目"
        ws["C4"] = "招标文件要求"
        ws["E4"] = "投标响应内容"
        ws["G4"] = "偏离情况"
        ws["H4"] = "说明"
        ws["A5"] = "承担《供货清单(一)》中各项物资生产组货任务"

        for ref in ["A2", "C2", "A3", "C3"]:
            ws[ref].font = self.header_font
            ws[ref].alignment = left_wrap

        for row in range(4, blank_row):
            for col in range(1, 9):
                cell = ws.cell(row, col)
                cell.font = self.normal_font
                cell.alignment = center_wrap
                cell.border = self.border

        for row in range(4, blank_row):
            for col in [3, 4, 5, 6]:
                ws.cell(row, col).alignment = left_wrap

        for idx, key in enumerate(sorted_keys):
            row_base = 6 + idx * 9
            item = items[key]
            item_name = item[0] if len(item) > 0 else ""
            hs_code = item[1] if len(item) > 1 else ""
            quantity = item[2] if len(item) > 2 else ""
            unit = item[3] if len(item) > 3 else ""
            spec_text = item[4] if len(item) > 4 else ""
            inspection = item[5] if len(item) > 5 else ""

            ws[f"A{row_base}"] = idx + 1
            ws[f"B{row_base}"] = f'="物资"&A{row_base}'

            for offset, label in enumerate(label_rows):
                ws[f"C{row_base + offset}"] = label
            for offset, label in enumerate(response_labels):
                ws[f"E{row_base + offset}"] = label

            ws[f"D{row_base}"] = item_name
            ws[f"D{row_base + 1}"] = f"{unit}{quantity}"
            ws[f"D{row_base + 2}"] = hs_code
            ws[f"D{row_base + 3}"] = spec_text
            ws[f"D{row_base + 4}"] = inspection
            ws[f"D{row_base + 5}"].border = self.diag_border
            ws[f"D{row_base + 6}"].border = self.diag_border
            ws[f"D{row_base + 7}"].border = self.diag_border
            ws[f"D{row_base + 8}"] = "无"

            ws[f"F{row_base}"] = f'=INDIRECT("物资选择!B"&A{row_base})'
            ws[f"F{row_base + 1}"] = (
                f'=INDEX(全部厂家备用!E$1:E${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
                f'&INDEX(全部厂家备用!D$1:D${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"F{row_base + 2}"] = (
                f'=INDEX(全部厂家备用!C$1:C${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"F{row_base + 3}"] = (
                f'="型号："&INDEX(全部厂家备用!I$1:I${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
                f'&CHAR(10)&INDEX(全部厂家备用!F$1:F${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"F{row_base + 4}"] = (
                f'=INDEX(全部厂家备用!G$1:G${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"F{row_base + 5}"] = (
                f'=INDEX(全部厂家备用!H$1:H${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"F{row_base + 6}"] = (
                f'=INDEX(全部厂家备用!L$1:L${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"F{row_base + 7}"] = (
                f'=INDEX(全部厂家备用!M$1:M${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"F{row_base + 8}"] = (
                f'=INDEX(全部厂家备用!O$1:O${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )
            ws[f"H{row_base + 3}"] = (
                f'=INDEX(全部厂家备用!AB$1:AB${supplier_last_row},INDIRECT("物资选择!C"&A{row_base}))'
            )

            for offset in range(9):
                ws[f"G{row_base + offset}"] = "无偏离"

            for ref in [f"D{row_base + 3}", f"F{row_base + 3}", f"H{row_base + 3}"]:
                ws[ref].font = small_font
                ws[ref].alignment = left_wrap

        ws[f"F{stamp_row}"] = "投标人盖章："
        ws[f"F{date_row}"] = "日期："
        ws[f"G{date_row}"] = self._parse_date(self.project.date)
        ws[f"F{stamp_row}"].font = footer_font
        ws[f"F{date_row}"].font = footer_font
        ws[f"G{date_row}"].font = footer_font
        ws[f"F{stamp_row}"].alignment = self.right
        ws[f"F{date_row}"].alignment = self.right
        ws[f"G{date_row}"].alignment = center_wrap
        ws[f"G{date_row}"].number_format = 'yyyy"年"m"月"d"日"'

    def _build_system_sheet(self, ws) -> None:
        ws.title = "16.三体系一览表"
        supplier_last_row = max(getattr(self, "_all_suppliers_last_row", 1), 1)
        main_items = list(getattr(self.project, "main_item", []) or [])

        self._set_columns(
            ws,
            {
                "A": 6.0,
                "B": 10,
                "C": 18,
                "D": 18,
                "E": 18,
                "F": 28,
                "G": 34.0,
                "H": 25,
                "I": 20,
            },
        )

        ws.merge_cells("A1:I1")
        ws.row_dimensions[1].height = 45
        ws.row_dimensions[2].height = 40

        ws["A1"] = "十七、主要标的的生产企业质量管理、环境管理和职业健康安全管理体系一览表"
        ws["A1"].font = Font(name="黑体", size=18)
        ws["A1"].alignment = self.center

        headers = [
            "序号",
            "物资序号",
            "物资名称",
            "品牌",
            "型号",
            "生产企业名称",
            "认证文件名称",
            "认证文件编号",
            "证书效期",
        ]
        for col, value in enumerate(headers, start=1):
            ws.cell(2, col, value)
            ws.cell(2, col).font = self.header_font
            ws.cell(2, col).alignment = self.center
            ws.cell(2, col).border = self.border

        cert_rows = [
            ("质量管理体系认证证书", "R", "S"),
            ("环境管理体系认证证书", "T", "U"),
            ("职业健康安全管理体系认证证书", "V", "W"),
        ]

        current_row = 3
        for item_no in main_items:
            for cert_name, cert_col, date_col in cert_rows:
                ws.row_dimensions[current_row].height = 30
                ws[f"A{current_row}"] = "=ROW()-2"
                ws[f"B{current_row}"] = item_no
                ws[f"C{current_row}"] = f'=INDIRECT("物资选择!B"&B{current_row})'
                ws[f"D{current_row}"] = (
                    f'=INDEX(全部厂家备用!H$1:H${supplier_last_row},INDIRECT("物资选择!C"&B{current_row}))'
                )
                ws[f"E{current_row}"] = (
                    f'=INDEX(全部厂家备用!I$1:I${supplier_last_row},INDIRECT("物资选择!C"&B{current_row}))'
                )
                ws[f"F{current_row}"] = (
                    f'=INDEX(全部厂家备用!L$1:L${supplier_last_row},INDIRECT("物资选择!C"&B{current_row}))'
                )
                ws[f"G{current_row}"] = cert_name
                ws[f"H{current_row}"] = (
                    f'=INDEX(全部厂家备用!{cert_col}$1:{cert_col}${supplier_last_row},INDIRECT("物资选择!C"&B{current_row}))'
                )
                ws[f"I{current_row}"] = (
                    f'=INDEX(全部厂家备用!{date_col}$1:{date_col}${supplier_last_row},INDIRECT("物资选择!C"&B{current_row}))'
                )
                ws[f"I{current_row}"].number_format = 'yyyy"年"m"月"d"日"'

                for col in range(1, 10):
                    cell = ws.cell(current_row, col)
                    cell.font = self.normal_font
                    cell.alignment = self.center
                    cell.border = self.border
                current_row += 1


    def generate(self, filename: Optional[str] = None) -> str:
        items = self.project.commodities
        bid_date = self._parse_date(self.project.date)

        wb = Workbook()
        ws_all = wb.active
        ws_fee = wb.create_sheet("运输费用")
        ws_other_fee = wb.create_sheet("其他费用")
        ws_open = wb.create_sheet("3.开标一览表")
        ws_procurement = wb.create_sheet("2.采购需求偏离表(物资部分)")
        ws_total = wb.create_sheet("1.投标报价总表")
        ws_inner = wb.create_sheet("2.物资对内分项报价表")
        ws_tax = wb.create_sheet("3.各项物资退抵税额表")
        ws_tech = wb.create_sheet("4.技术服务费报价表") if self.project.is_tech else None
        ws_train = wb.create_sheet("5.来华培训费报价表") if self.project.is_cc else None
        ws_system = wb.create_sheet("16.三体系一览表")
        ws_pick = wb.create_sheet("物资选择")

        self._build_all_suppliers(ws_all, items)
        self._build_selector(ws_pick, items)
        self._build_procurement_deviation_sheet(ws_procurement, items)
        self._build_fee_input(ws_fee, items)
        self._build_other_fees(ws_other_fee)
        inner_total_row = self._build_inner_quote(ws_inner, len(items))
        self._build_tax_sheet(ws_tax, len(items), inner_total_row)
        self._build_system_sheet(ws_system)
        if ws_tech is not None:
            self._build_tech_sheet(ws_tech)
        if ws_train is not None:
            self._build_train_sheet(ws_train)
        self._build_total_sheet(ws_total, inner_total_row, self._tax_total_row, bid_date)
        self._build_opening_sheet(ws_open, bid_date)

        wb.calculation.fullCalcOnLoad = True
        wb.calculation.iterate = True
        wb.calculation.iterateCount = 100
        wb.calculation.iterateDelta = 0.001
        if not filename:
            filename = f"投标报价表-{self._safe_name(self.project.name)}.xlsx"
        wb.save(filename)
        return filename
